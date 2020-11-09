from openpyxl import load_workbook

wb = load_workbook('18.Turnover.xlsx')
ws = wb.active

for row_no in range(2, ws.max_row + 1):
    ws.row_dimensions[row_no].hidden = False  # show hidden rows

for col_no in range(2, ws.max_column + 1):
    # get the column number from the cell
    col_alphabet = ws.cell(row=1, column=col_no).column_letter
    ws.column_dimensions[col_alphabet].hidden = False  # show hiden columns

wb.save('18.Turnover_changed.xlsx')
