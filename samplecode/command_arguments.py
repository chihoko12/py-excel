import sys

from openpyxl import load_workbook

# get the workbook name
filename = sys.argv[1]

# get the cell name
cellno = sys.argv[2]

wb = load_workbook(filename, read_only=True)
ws = wb.active

# run the progrmm by defining the filename and the cell, the value is shown on the console.
print(ws[cellno].value)
