from openpyxl import load_workbook

filename = input('the workbook to be opened: ')
cellno = input('the cell to be read(eg. A1): ')

wb = load_workbook(filename, read_only=True)
ws = wb.active

print(ws[cellno].value)
