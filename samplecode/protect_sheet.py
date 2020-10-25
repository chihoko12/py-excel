from openpyxl import load_workbook
from openpyxl.styles import Protection

wb = load_workbook('16.quote.xlsx')
ws = wb['見積書']

# unprotect the selected the cells
for rows in ws['B11:H24']:
    for cell in rows:
        cell.protection = Protection(locked=False)

# set the password for sheet protection
ws.protection.password = 'test'
# protect the sheet
ws.protection.enable()

wb.save('16.quote2_changed.xlsx')
