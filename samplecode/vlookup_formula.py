from openpyxl import load_workbook

wb = load_workbook('./books/21/21.workhour_lastmonth.xlsx')

lastmonth = '202004'
month = '202005'

ws_latmonth = wb[lastmonth]
ws = wb[month]

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row_count = row[0].row
    row[4].value = f'=VLOOKUP(B{row_count},{lastmonth}!$B$2:$D$11,3,FALSE)'

wb.save('./books/21/21.workhour_lastmonth_changed.xlsx')
