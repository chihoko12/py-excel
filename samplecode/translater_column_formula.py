from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

wb = load_workbook('./books/21/21.turnover_worker.xlsx')
ws = wb.active

ws['D3'] = Translator(ws['C3'].value, origin='C3').translate_formula('D3')

wb.save('./books/21/21.turnover_worker_changed.xlsx')
