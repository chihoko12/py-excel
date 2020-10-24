from openpyxl import load_workbook

wb = load_workbook('13.checklist.xlsx')

wb.move_sheet('まとめ', offset=len(wb.sheetnames))

wb.save('13.checklist_end_changed.xlsx')
