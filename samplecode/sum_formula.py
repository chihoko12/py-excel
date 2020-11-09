from openpyxl import load_workbook

wb = load_workbook('./books/21/21.workhour.xlsx')
ws = wb.active

ws['D2'] = '=SUM(D4:D13)'

wb.save('./books/21/21.workhour_changed.xlsx')
