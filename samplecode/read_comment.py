from openpyxl import load_workbook, Workbook
from openpyxl.comments import Comment

wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = 'comment list'

wb = load_workbook('20.schedule.xlsx')
ws = wb.active

ws_new['B2'] = 'comment description'
ws_new['C2'] = 'commented by'
ws_new['D2'] = 'cell no'

ws_new.column_dimensions['B'].width = 40
row_count = ws_new.max_row

for row in ws.iter_rows(min_row=4):
    for cell in row:
        if cell.comment is None:
            continue
        row_count = row_count + 1
        ws_new[f'B{row_count}'] = cell.comment.text
        ws_new[f'C{row_count}'] = cell.comment.author
        ws_new[f'D{row_count}'] = cell.coordinate

ws_new['D2'].comment = Comment('cell no which had comment', 'sato sachiko')
wb_new.save('20.comment_list.xlsx')
