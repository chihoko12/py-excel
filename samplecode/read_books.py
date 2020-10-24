from pathlib import Path
from openpyxl import load_workbook, Workbook

wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = 'List'

ws_new['B2'] = 'Department'
ws_new['C2'] = 'Name'

path = Path('././books')
for i, file in enumerate(path.glob('*.xlsx')):
    wb = load_workbook(file, read_only=True)
    ws = wb['checklist']

    row_no = i + 3
    ws_new[f'B{row_no}'] = ws['C2'].value
    ws_new[f'C{row_no}'] = ws['C3'].value

wb_new.save('List.xlsx')
