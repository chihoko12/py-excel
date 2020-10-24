from pathlib import Path
from openpyxl import load_workbook, Workbook

wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = 'summary'

ws_new['B2'] = 'Workbook name'
ws_new['C2'] = 'Total number of workbooks'
ws_new['D2'] = 'Total number of hidden workbooks'

path = Path('././books')
for i, file in enumerate(path.glob('*.xlsx')):
    wb = load_workbook(file)

    row_no = i + 3
    ws_new[f'B{row_no}'] = file.name
    ws_new[f'C{row_no}'] = len(wb.sheetnames)

    hidden_worksheets = [
        ws for ws in wb.worksheets if ws.sheet_state !=
        ws.SHEETSTATE_VISIBLE]
    ws_new[f'D{row_no}'] = len(hidden_worksheets)

wb_new.save('12.sheet_summary.xlsx')
