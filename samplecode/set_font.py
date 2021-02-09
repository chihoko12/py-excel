from openpyxl import load_workbook
from openpyxl.styles import Font

wb = load_workbook('25.sales_turnover.xlsx')
ws = wb.active
blue_font = Font(name='Arial', color='0000FF', size=18, bold=True)

for row in ws['B2':'F2']:
  for cell in row:
    cell.font = blue_font

wb.save('25.sales_turnover_changed.xlsx')
