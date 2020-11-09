import argparse

from openpyxl import load_workbook
# make argparse.ArgumentParser object
parser = argparse.ArgumentParser(
    description="Program to get the cell value from Excel")
# prepare the arguments
parser.add_argument('filename', help='the workbook to be open: ')
parser.add_argument('cellno', help='the cell no to be read (eg. A1): ')

# get the arguments
args = parser.parse_args()

wb = load_workbook(args.filename, read_only=True)
ws = wb.active

print(ws[args.cellno].value)
