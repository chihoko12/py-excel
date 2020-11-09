from openpyxl import load_workbook

wb = load_workbook('19.Turnover_per_item.xlsx')
ws = wb.active

num = 3
start_row = 3
for row_no in range(ws.max_row-num+1, num+start_row, -num):
    ws.insert_rows(row_no)
# insert rows from the bottom of the row

wb.save('19.Turnover_per_item_changed.xlsx')
